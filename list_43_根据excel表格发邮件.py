import openpyxl
import smtplib
import sys
from email.mime.text import MIMEText

wb = openpyxl.load_workbook(r'D:\Download\paid.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value
print('sheet.max_row is {}'.format(sheet.max_row))

unpaidMembers = {}

for r in range(2, sheet.max_row+1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=2, column=2).value
        if email != None:
            unpaidMembers[name] = email

smtpObj = smtplib.SMTP('smtp.163.com', 25)
smtpObj.starttls()
# smtpObj.login('ghfuidy@163.com', sys.argv[1])
smtpObj.login('ghfuidy@163.com', 'zyq1996110163')

for name, email in unpaidMembers.items():
    body = '\n我说那个{}啊,\n\n' \
           '{}的党费要交了！！！\n\n' \
           '大咖村党支部'.format(name, latestMonth)
    print('Sending email to {}...'.format(email))

    msg = MIMEText(body)
    msg['From'] = 'ghfuidy@163.com'
    msg['To'] = email
    msg['Subject'] = '拖欠的人可耻'
    sendmailStatus = smtpObj.send_message(msg)

    if sendmailStatus != {}:
        print('There was problem sending email to {}: {}'.format(
            email, sendmailStatus))
smtpObj.quit()

# import sendMessage
# message = '快点交党费？？？'
# sendMessage.textMyself(message)
